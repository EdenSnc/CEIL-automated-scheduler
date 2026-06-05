# data_manager.py
import json
import copy
import logging
import os
import ast
import sqlite3
import threading
import uuid
from typing import Any, Dict, List
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment

logger = logging.getLogger(__name__)

class ScheduleDataManager:
    _DEFAULT_METADATA: Dict[str, Any] = {
        "days": ["SAT", "SUN", "MON", "TUE", "WED", "THU"],
        "slots": [
            {"id": 0, "label": "08:30-10:30"},
            {"id": 1, "label": "10:30-12:30"},
            {"id": 2, "label": "13:30-15:30"},
            {"id": 3, "label": "15:30-17:30"},
            {"id": 4, "label": "17:30-19:30"},
        ],
        "solver_timeout": 60 # Added Execution Time State
    }

    _DEFAULT_WEIGHTS: Dict[str, int] = {
        "W_GAP": 30,
        "W_PREFERENCE": 1,
        "W_LOAD_BALANCE": 20,
        "W_FAIRNESS": 15,
        "W_ROOM_CHANGE": 10,
        "W_COGNITIVE": 5
        # W_SAME_DAY_CLUSTER removed entirely
    }

    def __init__(self, filepath: str = "ceil_scheduler.db") -> None:
        if filepath.endswith(".json"):
            filepath = filepath.replace(".json", ".db")
            
        if not os.path.isabs(filepath):
            base_dir = os.path.dirname(os.path.abspath(__file__))
            filepath = os.path.join(base_dir, filepath)
            
        self.filepath = filepath
        self.lock = threading.Lock()
        
        self.data: Dict[str, Any] = {
            "metadata": copy.deepcopy(self._DEFAULT_METADATA),
            "weights": copy.deepcopy(self._DEFAULT_WEIGHTS),
            "teachers": [],
            "groups": [],
            "rooms": [],
            "schedule": [],
            "locked_sessions": [], # Added support for Pinned/Makeup Sessions
        }
        self._init_db()
        self.load_from_disk()

    def _get_connection(self):
        conn = sqlite3.connect(self.filepath)
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        return conn

    def _init_db(self) -> None:
        try:
            with self._get_connection() as conn:
                conn.execute("""
                    CREATE TABLE IF NOT EXISTS application_state (
                        key TEXT PRIMARY KEY,
                        value TEXT
                    );
                """)
                conn.commit()
        except Exception as exc:
            logger.exception("Failed to initialize SQLite Database: %s", exc)

    def load_from_disk(self) -> None:
        try:
            with self._get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT key, value FROM application_state")
                rows = cursor.fetchall()
                
                if not rows:
                    return 
                    
                disk_payload = {}
                for key, val_str in rows:
                    disk_payload[key] = json.loads(val_str)
                
                for key in ("metadata", "teachers", "groups", "rooms", "schedule", "locked_sessions"):
                    if key in disk_payload:
                        self.data[key] = disk_payload[key]
                
                if "weights" in disk_payload:
                    for k, v in disk_payload["weights"].items():
                        if k != "W_SAME_DAY_CLUSTER": # Filter out legacy key if present on disk
                            self.data["weights"][k] = v
        except Exception as exc:
            logger.exception("Error loading state from SQLite: %s", exc)

    def save_to_disk(self) -> None:
        state_snapshot = copy.deepcopy(self.data)
        threading.Thread(target=self._execute_db_write, args=(state_snapshot,), daemon=True).start()

    def _execute_db_write(self, state_snapshot: Dict[str, Any]) -> None:
        with self.lock:
            try:
                with self._get_connection() as conn:
                    for key in ["metadata", "weights", "teachers", "groups", "rooms", "schedule", "locked_sessions"]:
                        payload_str = json.dumps(state_snapshot.get(key, []), ensure_ascii=False)
                        conn.execute("""
                            INSERT INTO application_state (key, value) 
                            VALUES (?, ?)
                            ON CONFLICT(key) DO UPDATE SET value=excluded.value;
                        """, (key, payload_str))
                    conn.commit()
            except Exception as exc:
                logger.error("Background SQLite save pipeline dropped a write: %s", exc)

    # ==========================================
    # --- PINNED / MAKEUP SESSION MANAGEMENT ---
    # ==========================================

    def add_locked_session(self, group_id: str, teacher_id: str, room_id: str, day_idx: int, slot_id: str) -> None:
        if "locked_sessions" not in self.data:
            self.data["locked_sessions"] = []
            
        session_data = {
            "id": str(uuid.uuid4()),
            "group_id": group_id,
            "teacher_id": teacher_id,
            "room_id": room_id,
            "day_idx": day_idx,
            "slot_id": slot_id
        }
        self.data["locked_sessions"].append(session_data)
        self.save_to_disk()

    def get_locked_sessions(self) -> list:
        return self.data.get("locked_sessions", [])

    def remove_locked_session(self, session_id: str) -> None:
        if "locked_sessions" in self.data:
            original_len = len(self.data["locked_sessions"])
            self.data["locked_sessions"] = [
                s for s in self.data["locked_sessions"] if s.get("id") != session_id
            ]
            if len(self.data["locked_sessions"]) < original_len:
                self.save_to_disk()

    # ==========================================
    # --- STANDARD ENTITY MANAGEMENT ---
    # ==========================================

    def update_weight(self, key: str, value: int) -> None:
        self.data["weights"][key] = value
        self.save_to_disk()

    def reset_weights(self) -> None:
        self.data["weights"] = copy.deepcopy(self._DEFAULT_WEIGHTS)
        self.save_to_disk()

    _ID_PREFIX: Dict[str, str] = {"teachers": "T_", "groups": "G_", "rooms": "R_"}

    def _next_id(self, collection: str) -> str:
        prefix = self._ID_PREFIX[collection]
        indices: List[int] = []
        for item in self.data[collection]:
            raw = str(item.get("id", ""))
            if raw.startswith(prefix):
                try:
                    indices.append(int(raw[len(prefix):]))
                except ValueError:
                    pass
        return f"{prefix}{max(indices) + 1 if indices else 1}"

    def add_entity(self, collection: str, entity: Dict[str, Any]) -> None:
        if collection not in self._ID_PREFIX:
            return
        entity["id"] = self._next_id(collection)
        self.data[collection].append(entity)
        self.save_to_disk()

    def update_entity(self, collection: str, index: int, updated: Dict[str, Any]) -> None:
        if collection not in self._ID_PREFIX:
            return
        items = self.data[collection]
        if not (0 <= index < len(items)):
            return
        updated["id"] = items[index]["id"]
        items[index].update(updated)
        self.save_to_disk()

    def delete_entity(self, collection: str, index: int) -> None:
        if collection not in self._ID_PREFIX:
            return
        items = self.data[collection]
        if not (0 <= index < len(items)):
            return
        items.pop(index)
        self.save_to_disk()

    # ==========================================
    # --- EXCEL I/O PIPELINE ---
    # ==========================================

    def export_to_excel(self, filepath: str) -> None:
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            import pandas as pd
            import copy
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                schedule = self.data.get("schedule", [])
                
                # 1. 🟢 VISUAL TIMETABLES: ONE SHEET PER DAY (The Beautiful UI Grids)
                if schedule:
                    day_order = ["SAT", "SUN", "MON", "TUE", "WED", "THU"]
                    lang_colors = {
                        "english": "C0DAFE", "french": "FB70BF", "spanish": "FFB554",
                        "italian": "52E084", "german": "907C63", "swedish": "86F5DD",
                        "japanese": "F9CBCB", "chinese": "F97171", "turkish": "C192F4",
                        "russian": "6E6EFB"
                    }
                    
                    rooms = [r.get("name", r.get("id")) for r in self.data.get("rooms", [])]
                    slots_metadata = self.data.get("metadata", {}).get("slots", [])
                    slots = [s["label"] for s in slots_metadata] if slots_metadata else ["08:30-10:30", "10:30-12:30", "13:30-15:30", "15:30-17:30", "17:30-19:30"]

                    workbook = writer.book
                    border_side = Side(style="thin", color="CBD5E1")
                    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
                    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
                    time_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

                    active_days = set(i.get("day", "").upper() for i in schedule)
                    
                    for day in day_order:
                        if day not in active_days: continue
                        
                        day_items = [i for i in schedule if i.get("day", "").upper() == day]
                        ws = workbook.create_sheet(f"{day} Timetable")
                        
                        ws.cell(row=1, column=1, value="Time Slot").font = header_font
                        ws.cell(row=1, column=1).fill = header_fill
                        ws.cell(row=1, column=1).border = cell_border
                        ws.column_dimensions['A'].width = 16
                        
                        for col_idx, room in enumerate(rooms, start=2):
                            cell = ws.cell(row=1, column=col_idx, value=room)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.border = cell_border
                            ws.column_dimensions[cell.column_letter].width = 24
                            
                        for row_idx, slot in enumerate(slots, start=2):
                            t_cell = ws.cell(row=row_idx, column=1, value=slot)
                            t_cell.font = Font(name="Segoe UI", size=10, bold=True, color="475569")
                            t_cell.fill = time_fill
                            t_cell.border = cell_border
                            t_cell.alignment = Alignment(horizontal="center", vertical="center")
                            ws.row_dimensions[row_idx].height = 45
                            
                            for col_idx, room in enumerate(rooms, start=2):
                                match = next((i for i in day_items if i.get("slot") == slot and i.get("room") == room), None)
                                cell = ws.cell(row=row_idx, column=col_idx)
                                cell.border = cell_border
                                
                                if match:
                                    grp = match.get("group", "")
                                    tch = match.get("teacher", "")
                                    lang = str(match.get("language", "english")).lower()
                                    
                                    cell.value = f"{grp}\n({tch})"
                                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
                                    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                                    
                                    color_hex = lang_colors.get(lang, "E2E8F0")
                                    cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

                # 2. 🟢 THE "EVERYTHING TABLE" (Master Filterable Schedule)
                if schedule:
                    schedule_df = pd.DataFrame(schedule)
                    # Reorder columns so the most important data is read left-to-right
                    preferred_cols = ["day", "slot", "room", "group", "teacher", "language"]
                    existing_cols = [c for c in preferred_cols if c in schedule_df.columns]
                    # Add any remaining random columns to the end
                    existing_cols += [c for c in schedule_df.columns if c not in preferred_cols]
                    
                    schedule_df = schedule_df[existing_cols]
                    # Rename columns to look professional
                    schedule_df.columns = [str(c).title() for c in schedule_df.columns]
                    schedule_df.to_excel(writer, sheet_name="Master Schedule", index=False)

                # 3. PROTECT PINNED MAKEUP EXCEPTIONS IN THE WORKBOOK
                locked = self.data.get("locked_sessions", [])
                if locked:
                    pd.DataFrame(locked).to_excel(writer, sheet_name="Pinned Exceptions", index=False)

                # 4. MASTER ENTITY VIEWS
                teachers_copy = copy.deepcopy(self.data.get("teachers", []))
                for t in teachers_copy:
                    t.pop("allowed_days", None)
                    if "allowed_slots" in t and isinstance(t["allowed_slots"], list):
                        t["allowed_slots"] = ", ".join(str(s) for s in t["allowed_slots"])
                    if "preferences" in t:
                        t["preferences"] = str(t["preferences"])
                if teachers_copy:
                    pd.DataFrame(teachers_copy).to_excel(writer, sheet_name="Teachers", index=False)

                groups_copy = copy.deepcopy(self.data.get("groups", []))
                for g in groups_copy:
                    g.pop("allowed_days", None)
                    if "allowed_slots" in g and isinstance(g["allowed_slots"], list):
                        g["allowed_slots"] = ", ".join(str(s) for s in g["allowed_slots"])
                if groups_copy:
                    pd.DataFrame(groups_copy).to_excel(writer, sheet_name="Groups", index=False)
                
                if self.data.get("rooms"):
                    pd.DataFrame(self.data.get("rooms", [])).to_excel(writer, sheet_name="Rooms", index=False)

                # Clean up empty default sheet
                if "Sheet" in writer.book.sheetnames:
                    del writer.book["Sheet"]
                    
                global_font = Font(name="Segoe UI", size=11)
                header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
                empty_fill  = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                
                for sheet_name in writer.book.sheetnames:
                    if "Timetable" in sheet_name:
                        continue # Leave our custom visual grids alone
                    
                    ws = writer.book[sheet_name]
                    
                    # 🟢 Turn on Excel Auto-Filters for the Master Everything Table
                    if sheet_name == "Master Schedule":
                        ws.auto_filter.ref = ws.dimensions
                        ws.freeze_panes = "A2" # Freeze the top header row so it follows you when scrolling down!
                        
                    for row_idx, row in enumerate(ws.iter_rows(), start=1):
                        is_header = (row_idx == 1)
                        if is_header:
                            ws.row_dimensions[row_idx].height = 26
                        else:
                            has_newline = any(str(cell.value).find('\n') != -1 for cell in row if cell.value)
                            if not has_newline:
                                ws.row_dimensions[row_idx].height = 20
                                
                        for cell in row:
                            if is_header:
                                cell.font = header_font
                                cell.fill = header_fill
                            else:
                                cell.font = global_font
                                cell.fill = empty_fill
                            
                            if cell.value:
                                lines = str(cell.value).split('\n')
                                max_length = max(len(line) for line in lines)
                                current_width = ws.column_dimensions[cell.column_letter].width or 12
                                ws.column_dimensions[cell.column_letter].width = max(max_length + 4, current_width)

        except Exception as e:
            logger.error(f"Excel export configuration crashed: {e}")
            raise

    def import_master_data_from_excel(self, filepath: str) -> None:
        try:
            xls = pd.ExcelFile(filepath)
            
            # 🟢 SAFE IMPORT PIPELINE: Looks for our new Master sheet, but falls back to older legacy names automatically
            target_schedule_sheet = None
            for legacy_name in ["Master Schedule", "Raw Schedule", "Schedule"]:
                if legacy_name in xls.sheet_names:
                    target_schedule_sheet = legacy_name
                    break
                    
            if target_schedule_sheet:
                sched_df = pd.read_excel(xls, target_schedule_sheet)
                sched_df = sched_df.where(pd.notnull(sched_df), None)
                
                # Normalize columns back to internal lowercase keys (Day, Slot -> day, slot)
                sched_df.columns = [str(c).lower() for c in sched_df.columns]
                
                raw_assignments = sched_df.to_dict(orient="records")
                
                clean_assignments = []
                for a in raw_assignments:
                    clean_a = {k: v for k, v in a.items() if v is not None}
                    if clean_a:
                        clean_assignments.append(clean_a)
                
                self.data["schedule"] = clean_assignments
                logger.info(f"[Import Pipeline] Successfully synchronized {len(clean_assignments)} rows from '{target_schedule_sheet}'.")

            # RESTORE YOUR LOCKED MAKEUP SESSIONS FROM THE SPREADSHEET
            if "Pinned Exceptions" in xls.sheet_names:
                locked_df = pd.read_excel(xls, "Pinned Exceptions")
                locked_df = locked_df.where(pd.notnull(locked_df), None)
                raw_locked = locked_df.to_dict(orient="records")
                
                clean_locked = []
                for l in raw_locked:
                    clean_l = {k: v for k, v in l.items() if v is not None}
                    if clean_l:
                        clean_locked.append(clean_l)
                        
                self.data["locked_sessions"] = clean_locked
                logger.info(f"[Import Pipeline] Restored {len(clean_locked)} pinned exceptions from backup.")

            def parse_target_sheet(sheet_name):
                if sheet_name in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name)
                    df = df.where(pd.notnull(df), None)
                    records = df.to_dict(orient="records")
                    cleaned_records = []
                    for rec in records:
                        clean_r = {k: v for k, v in rec.items() if v is not None}

                        if "language" in rec:
                            raw_language = rec.get("language")
                            clean_r["language"] = "" if raw_language is None else str(raw_language)
                        
                        clean_r.pop("allowed_days", None)
                        
                        if "allowed_slots" in clean_r:
                            val_str = str(clean_r["allowed_slots"]).replace("[", "").replace("]", "").replace("'", "")
                            if val_str.strip() == "":
                                clean_r["allowed_slots"] = []
                            else:
                                tokens = [t.strip() for t in val_str.split(",")]
                                int_slots = []
                                for t in tokens:
                                    try:
                                        int_slots.append(int(float(t)))
                                    except ValueError:
                                        pass
                                clean_r["allowed_slots"] = int_slots

                        if "preferences" in clean_r and isinstance(clean_r["preferences"], str):
                            try:
                                import ast
                                clean_r["preferences"] = ast.literal_eval(clean_r["preferences"])
                            except Exception:
                                pass 
                        cleaned_records.append(clean_r)
                    return cleaned_records
                return []

            teachers = parse_target_sheet("Teachers")
            groups = parse_target_sheet("Groups")
            rooms = parse_target_sheet("Rooms")
            
            if teachers:
                for t in teachers: t.pop("allowed_days", None)
                self.data["teachers"] = teachers
            if groups:
                for g in groups: g.pop("allowed_days", None)
                self.data["groups"] = groups
            if rooms: 
                self.data["rooms"] = rooms
            
            self.save_to_disk()
            logger.info("[Import Pipeline] Complete workbook synchronization routine executed successfully.")
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"Excel import failed: {e}")
            raise